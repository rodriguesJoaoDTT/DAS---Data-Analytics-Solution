# Library import

import pandas as pd
from common.base_path import get_base_dir
import numpy as np
from datetime import datetime
import os
import shutil
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from common.version import __DASversion__, __TARAversion__
from pathlib import Path

# File import Def

def files_import(input_path):

    BASE_DIR = get_base_dir()
        
    sheets = [
        'general_summary',
        'system_summary',
        'ativos',
        'desligados',
        'usuarios_atributo_a',
        'usuarios_atributo_b'
    ]

    all_sheets = pd.read_excel(input_path, sheet_name=sheets)

    dataframes = {f"df_{sheet}": df for sheet, df in all_sheets.items()}

    return dataframes, BASE_DIR, input_path


# Registro de Qtds

def qtd_register(df1, df2):
    # Contagem por sistema em cada df
    contagem_1 = df1['Sistema'].value_counts()
    contagem_2 = df2['Sistema'].value_counts()

    # Soma os valores, unindo os dois índices
    contagem_total = contagem_1.add(contagem_2, fill_value=0).astype(int)

    # Transforma em DataFrame
    df_resultado = contagem_total.reset_index()
    df_resultado.columns = ['Sistema', 'Qtd']

    return df_resultado

# ID normalize Defs

def id_normalize(x):
    try:
        return x.strip().upper()
    except AttributeError:
        return x

def id_normalize_apply(df, coluna='ID'):
    df[coluna] = df[coluna].apply(id_normalize)
    return df


# FY Manipulation Defs

def fy_dates(df_summary, col_info='Informação', col_valor='Valor'):
    fy_start = pd.to_datetime(df_summary.loc[df_summary[col_info] == 'FY Start', col_valor].values[0])
    fy_end   = pd.to_datetime(df_summary.loc[df_summary[col_info] == 'FY End', col_valor].values[0])
    return fy_start, fy_end

def fy_filter(df_desligados, fy_start, fy_end, col_data='Data de Desligamento'):
    df = df_desligados.copy()
    df[col_data] = pd.to_datetime(df[col_data], errors='coerce')
    df_filtrado = df[(df[col_data] >= fy_start) & (df[col_data] <= fy_end)].copy()
    return df_filtrado

# False Positive From Actives Remove

def actives_false_positive(df_desligados_fy, df_ativos):

    qtd_antes = df_desligados_fy.shape[0]
    
    df_filtrado = df_desligados_fy[~df_desligados_fy['ID'].isin(df_ativos['ID'])].copy()
    
    qtd_depois = df_filtrado.shape[0]
    qtd_falsos_positivos = qtd_antes - qtd_depois

    return df_filtrado, qtd_falsos_positivos


# Pivot Tables e Df de datas de bloqueio

def pivot_attribute_a(df):

    df = df.drop_duplicates(subset=['ID', 'Sistema'])
    return df.pivot_table(index='ID', columns='Sistema', values='Status', aggfunc='first').reset_index()

def pivot_attribute_b(df):
  
    df = df.drop(['Data de Bloqueio'], axis=1)
    df = df.drop_duplicates(subset=['ID', 'Sistema'])
    return df.pivot_table(index='ID', columns='Sistema', values='Status', aggfunc='first').reset_index()

def attribute_b_date(df):
  
    return df.drop_duplicates(subset=['ID', 'Sistema'])


# Teste do atributo A

def df_test_attribute_a(df_pivot_a, df_pivot_b):
    
    return pd.merge(df_pivot_a, df_pivot_b, on='ID', how='outer')


def df_desligados_access(df_desligados_fy, df_teste_atributo_a):
    
    return df_desligados_fy.merge(df_teste_atributo_a, on='ID', how='left')


def df_pop_attribute_a(df_desligados_acessos):
   
    colunas_originais = ['ID', 'Nome', 'Cargo', 'Centro de Custo', 'Data de Desligamento']
    colunas_sistemas = [col for col in df_desligados_acessos.columns if col not in colunas_originais]

    df_pop = df_desligados_acessos.dropna(subset=colunas_sistemas, how='all').copy()
    sem_acesso = len(df_desligados_acessos) - len(df_pop)
    
    df_pop.loc[:, colunas_sistemas] = df_pop[colunas_sistemas].fillna("Sem Acesso")
    populacao = len(df_pop)

    return df_pop, sem_acesso, populacao, colunas_sistemas


def testing_attribute_a(df_populacao):

    df = df_populacao.copy()

    colunas_fixas = ['ID', 'Nome', 'Cargo', 'Centro de Custo', 'Data de Desligamento']
    colunas_dinamicas = [col for col in df.columns if col not in colunas_fixas]

    df.loc[:, 'Atributo A'] = df[colunas_dinamicas].apply(
        lambda row: 'Não' if (row == 'Ativo').any() else 'Sim',
        axis=1
    )

    df.loc[:, 'Conclusão'] = df['Atributo A'].map({
        'Sim': 'Effective',
        'Não': 'Ineffective'
    })

    return df


# Teste do Atributo B

def df_test_attribute_b(df_desligados_fy, df_atributo_b_data):
    
    return df_desligados_fy.merge(df_atributo_b_data, how='left', on='ID')


def df_pop_attribute_b(df_desligados_acessos_b):

    colunas_originais = ['ID', 'Nome', 'Cargo', 'Centro de Custo', 'Data de Desligamento']
    colunas_sistemas = [col for col in df_desligados_acessos_b.columns if col not in colunas_originais]

    df_pop = df_desligados_acessos_b.dropna(subset=colunas_sistemas, how='all').copy()
    return df_pop

def testing_attribute_b(df_populacao, df_system_summary):
    
    df = df_populacao.copy()

    df['Data de Bloqueio'] = pd.to_datetime(df['Data de Bloqueio'], errors='coerce')
    df['Data de Desligamento'] = pd.to_datetime(df['Data de Desligamento'], errors='coerce')

    df['Tempestividade'] = (df['Data de Bloqueio'] - df['Data de Desligamento']).dt.days

    limites_tempestividade = df_system_summary.set_index('Sistema')['Tempestividade'].to_dict()

    def verifica_atributo_b(row):
        limite = limites_tempestividade.get(row['Sistema'], None)
        if limite is None or pd.isna(row['Tempestividade']):
            return None
        return 'Sim' if row['Tempestividade'] <= limite and row['Status'] == 'Bloqueado' else 'Não'

    df['Atributo B'] = df.apply(verifica_atributo_b, axis=1)

    df['Conclusão'] = df['Atributo B'].map({
        'Sim': 'Effective',
        'Não': 'Ineffective'
    })

    return df


# Formatar DFs de Saída

def cap_columns(df, colunas):
   
    df = df.copy()
    for col in colunas:
        if col in df.columns:
            df[col] = df[col].astype(str).str.title().where(df[col].notnull(), None)
    return df


def format_df_attribute_a(df):
    
    df = df.copy()
    df["Data de Desligamento"] = pd.to_datetime(df["Data de Desligamento"], errors='coerce')
    df = df.sort_values(by=["Data de Desligamento", "Nome"])
    df["Data de Desligamento"] = df["Data de Desligamento"].dt.strftime('%d/%m/%Y')
    return df


def format_df_attribute_b(df):
    
    df = df.copy()
    df["Data de Desligamento"] = pd.to_datetime(df["Data de Desligamento"], errors='coerce')
    df["Data de Bloqueio"] = pd.to_datetime(df["Data de Bloqueio"], errors='coerce')
    df = df.sort_values(by=["Data de Desligamento", "Nome", "Sistema"])
    df["Data de Desligamento"] = df["Data de Desligamento"].dt.strftime('%d/%m/%Y')
    df["Data de Bloqueio"] = df["Data de Bloqueio"].dt.strftime('%d/%m/%Y')
    return df


# DF Summary Attribute A

def df_summary_atributo_a(
    qtd_desligados_original,
    qtd_desligados_fy,
    qtd_falsos_positivos_ativos,
    sem_acesso,
    populacao,
    df_populacao_teste,
    colunas_sistemas
):
   
    
    dados_iniciais = [
        ["Terminated Employees (Original Data)", qtd_desligados_original, "", ""],
        ["FY Terminated Employees", qtd_desligados_fy, "", ""],
        ["Terminated Employee on Active Employees Report (False Positive)", qtd_falsos_positivos_ativos, "", ""],
        ["Terminated Employees Without Access", sem_acesso, "", ""],
        ["Terminate Employees Full Population", populacao, "", ""]
    ]

    df_resumo = pd.DataFrame(dados_iniciais, columns=["Information", "Value", "Attribute A", "Conclusion"])

    
    for coluna in colunas_sistemas:
        if coluna in df_populacao_teste.columns:
            qtd = df_populacao_teste[coluna].isin(["Ativo", "Bloqueado"]).sum()
            nova_linha = pd.DataFrame(
                [[f"Terminate Employees {coluna} Population", qtd, "", ""]],
                columns=df_resumo.columns
            )
            df_resumo = pd.concat([df_resumo, nova_linha], ignore_index=True)

    return df_resumo

# DF Extraction Date

def gera_df_extraction_date(df_general_summary, df_system_summary, colunas_sistemas):

    linhas = []

    
    data_ativos = df_general_summary.loc[
        df_general_summary["Informação"] == "Data de Extração Ativos", "Valor"
    ].values[0]
    linhas.append(["Active Employees", data_ativos])

    
    data_desligados = df_general_summary.loc[
        df_general_summary["Informação"] == "Data de Extração Desligados", "Valor"
    ].values[0]
    linhas.append(["Terminated Employees", data_desligados])

    
    for sistema in colunas_sistemas:
        if sistema in df_system_summary["Sistema"].values:
            data_extracao = df_system_summary.loc[
                df_system_summary["Sistema"] == sistema, "Data de Extração"
            ].values[0]
        else:
            data_extracao = "Não encontrado"

        linhas.append([sistema, data_extracao])

    
    df_extraction_date = pd.DataFrame(linhas, columns=["Fonte", "Data de Extração"])

   
    df_extraction_date["Data de Extração"] = df_extraction_date["Data de Extração"].apply(
        lambda x: pd.to_datetime(x).strftime('%d/%m/%Y') 
        if isinstance(x, (pd.Timestamp, datetime, np.datetime64)) 
        else str(x)
    )

    return df_extraction_date



def gera_df_qtds(qtd_desligados_original, qtd_ativos_original, qtd_por_sistema, colunas_sistemas):

    linhas = []

    linhas.append(["Active Employees", qtd_ativos_original])   
    linhas.append(["Terminated Employees", qtd_desligados_original])

    
    for sistema in colunas_sistemas:
        if sistema in qtd_por_sistema["Sistema"].values:
            qtd = qtd_por_sistema.loc[
                qtd_por_sistema["Sistema"] == sistema, "Qtd"
            ].values[0]
        else:
            qtd = "Não encontrado"

        linhas.append([sistema, qtd])

    
    df_quantidade = pd.DataFrame(linhas, columns=["Sistema", "Qtd"])

    return df_quantidade


# DF Tipo Teste

def gera_df_tipo_teste(df_system_summary, colunas_sistemas):
 

    linhas_teste = []

    for sistema in colunas_sistemas:
        if sistema in df_system_summary["Sistema"].values:
            tipo_teste = df_system_summary.loc[
                df_system_summary["Sistema"] == sistema, "Tipo de Teste"
            ].values[0]
        else:
            tipo_teste = "Não encontrado"

        linhas_teste.append([sistema, tipo_teste])

    df_tipo_teste = pd.DataFrame(linhas_teste, columns=["Sistema", "Tipo de Teste"])

    
    for col in df_tipo_teste.columns:
        df_tipo_teste[col] = df_tipo_teste[col].map(
            lambda x: pd.to_datetime(x).to_pydatetime() if isinstance(x, np.datetime64) else x
        )

    return df_tipo_teste


# DF Tempestividade

def df_tempestividade(df_system_summary):

    
    colunas_sistemas_ab = df_system_summary.loc[
        df_system_summary["Tipo de Teste"] == "Atributo A e B", "Sistema"
    ].tolist()

    linhas_temp = []

    for sistema in colunas_sistemas_ab:
        if sistema in df_system_summary["Sistema"].values:
            temp_value = df_system_summary.loc[
                df_system_summary["Sistema"] == sistema, "Tempestividade"
            ].values[0]
        else:
            temp_value = "N/A"

        linhas_temp.append([sistema, temp_value])

    df_temp_value = pd.DataFrame(linhas_temp, columns=["Sistema", "Tempestividade"])

    
    df_temp_value["Tempestividade"] = df_temp_value["Tempestividade"].apply(
        lambda x: str(int(x)) if isinstance(x, (int, float)) else str(x)
    )

    return df_temp_value


# Gerar arquivo base para output

def gerar_arquivos_output(df_general_summary, BASE_DIR, input_path):
    
    pasta_origem = BASE_DIR / 'data'
    pasta_destino = BASE_DIR / 'output'

    arquivo_origem = 'Working Paper Revogação de Acessos (Output) - Modelo.xlsx'

    nome_cliente = df_general_summary.loc[
        df_general_summary['Informação'] == 'Nome do Engagement', 'Texto'
    ].values[0]

    arquivo_copia = f"{nome_cliente} - GITC.02 - Revogação de Acessos.xlsx"
    input_path_out = f"{nome_cliente} - Template de Revogação de Acessos - X Sistemas (Input).xlsx"

    caminho_origem = os.path.join(pasta_origem, arquivo_origem)
    caminho_destino = os.path.join(pasta_destino, arquivo_copia)
    caminho_input = os.path.join(pasta_origem, input_path)
    caminho_input_out = os.path.join(pasta_destino, input_path_out)

    shutil.copy2(caminho_origem, caminho_destino)
    shutil.copy2(caminho_input, caminho_input_out)

    return {
        "caminho_output": caminho_destino,
        "nome_cliente": nome_cliente
    }


# Def para configurar o preenchimento do excel

# ---------- FUNÇÃO AUXILIAR PARA ESCREVER DF COM FORMATAÇÃO ----------
def escreve_df(ws, df, start_row, start_col, cores_sistemas):
    """Escreve df em ws a partir de (start_row, start_col) com formatação"""
    font_hdr   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_std   = Font(name="Calibri", size=11)
    font_eff   = Font(name="Calibri", size=11, bold=True, color="00B050")  # verde
    font_ineff = Font(name="Calibri", size=11, bold=True, color="FF0000")  # vermelho

    fill_preto       = PatternFill("solid", fgColor="000000")
    fill_azul_claro  = PatternFill("solid", fgColor="00B0F0")
    fill_verde_claro = PatternFill("solid", fgColor="92D050")
    fill_roxo        = PatternFill("solid", fgColor="7030A0")
    fill_azul_escuro = PatternFill("solid", fgColor="1F4E78")

    thin = Side(style="thin")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    cols_estaticas = ['ID', 'Nome', 'Cargo', 'Centro de Custo', 'Data de Desligamento']
    cols_sistemas  = [c for c in df.columns if c not in cols_estaticas + ['Atributo A', 'Atributo B', 'Tempestividade', 'Sistema', 'Status', 'Data de Bloqueio', 'Conclusão']]
    alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Cabeçalhos
    for idx, col in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + idx, value=col)
        cell.font = font_hdr
        cell.border = borda
        cell.alignment = alinhamento_centro

        if col in cols_estaticas:
            cell.fill = fill_preto
        elif col in ['Atributo A', 'Atributo B']:
            cell.fill = fill_azul_claro
        elif col == 'Conclusão':
            cell.fill = fill_verde_claro
        elif col == 'Tempestividade':
            cell.fill = fill_roxo
        elif col in ['Sistema','Status', 'Data de Bloqueio']:
            cell.fill = fill_azul_escuro
        else:  # colunas de sistemas
            cor = cores_sistemas[cols_sistemas.index(col) % len(cores_sistemas)]
            cell.fill = PatternFill("solid", fgColor=cor)

    # Dados
    for r, (_, row) in enumerate(df.iterrows(), start=1):
        for c, value in enumerate(row, start=0):
            cell = ws.cell(row=start_row + r, column=start_col + c, value=value)
            cell.font = font_std
            cell.border = borda
            cell.alignment = alinhamento_centro

            # Formatação especial da coluna Conclusão
            if df.columns[c] == 'Conclusão':
                if value == 'Effective':
                    cell.font = font_eff
                elif value == 'Ineffective':
                    cell.font = font_ineff


# ---------- FUNÇÃO PRINCIPAL PARA GERAR E FORMATAR O RELATÓRIO ----------
def escreve_df_formatado_em_excel(wb, df_teste_atributo_a, df_teste_atributo_b):
    # Paleta fixa p/ sistemas (sempre nesta ordem)
    cores_sistemas = [
        "1F4E78", "F79646", "548235", "0C4C8A", "C00000",
        "4BACC6", "9BBB59", "2F5597", "843C0C", "264478",
        "7F6000", "512E5F", "1B5E20", "7B241C", "0B4C5F"
    ]

    # Abrir planilha e limpar área a partir da linha_clear
    
    aba_destino = "APP.02 Test"
    linha_clear = 14
    linha_inicio_a = 15
    coluna_inicio = 2
    ws = wb[aba_destino]

    max_col_letter = get_column_letter(ws.max_column or 1)
    max_row = ws.max_row or linha_clear

    for row in ws[f"A{linha_clear}:{max_col_letter}{max_row}"]:
        for cell in row:
            cell.value = None
            cell.fill  = PatternFill()  # remove cor
            cell.font  = Font()         # remove fonte
            cell.border = Border()      # remove borda

    # Escrever df_teste_atributo_a
    ws.cell(row=linha_inicio_a - 2, column=coluna_inicio, value="Tabela 1 - Teste Atributo A").font = Font(name="Calibri", size=11, bold=True)
    escreve_df(ws, df_teste_atributo_a, start_row=linha_inicio_a, start_col=coluna_inicio, cores_sistemas=cores_sistemas)

    # Escrever df_teste_atributo_b logo abaixo
    linha_inicio_b = linha_inicio_a + len(df_teste_atributo_a) + 4
    ws.cell(row=linha_inicio_b - 2, column=coluna_inicio, value="Tabela 2 - Teste Atributo B").font = Font(name="Calibri", size=11, bold=True)
    escreve_df(ws, df_teste_atributo_b, start_row=linha_inicio_b, start_col=coluna_inicio, cores_sistemas=cores_sistemas)

    # Ajustar largura das colunas
    larguras_colunas = {
        "ID": 25,
        "Nome": 50,
        "Cargo": 50,
        "Centro de Custo": 50,
        "Data de Desligamento": 25,
    }
    largura_padrao = 20

    for idx, col_name in enumerate(df_teste_atributo_a.columns):
        col_letter = get_column_letter(coluna_inicio + idx)
        largura = larguras_colunas.get(col_name, largura_padrao)
        ws.column_dimensions[col_letter].width = largura


# Escrever a aba de Summary:

def escrever_summary(
    wb,
    nome_cliente,
    fy_start,
    df_summary_attribute_a,
    df_extraction_date,
    df_quantidade,
    df_tipo_teste,
    df_temp_value
):
    # Configurações básicas
    aba_destino_summary = "APP.02 Summary"
    linha_clear_summary = 45          # limpar a partir daqui
    linha_inicio_summary = 44         # onde começa o resumo
    coluna_inicio_summary = 2         # coluna B

    # Estilos - reutilizando os mesmos do contexto anterior
    font_hdr   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_std   = Font(name="Calibri", size=11)
    font_eff   = Font(name="Calibri", size=11, bold=True, color="00B050")  # verde
    font_ineff = Font(name="Calibri", size=11, bold=True, color="FF0000")  # vermelho
    fill_preto       = PatternFill("solid", fgColor="000000")
    fill_azul_claro  = PatternFill("solid", fgColor="00B0F0")
    fill_verde_claro = PatternFill("solid", fgColor="92D050")
    borda_preta = Border(left=Side(style="thin", color="000000"),
                         right=Side(style="thin", color="000000"),
                         top=Side(style="thin", color="000000"),
                         bottom=Side(style="thin", color="000000"))

    alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Função para escrever o df_summary_attribute_a (com cabeçalho e formatação)
    def escreve_df_summary(wss, df, start_row, start_col):
        cols_estaticas = ['Information', 'Value']
        
        # Cabeçalhos
        for idx, col in enumerate(df.columns):
            cell = wss.cell(row=start_row, column=start_col + idx, value=col)
            cell.font = font_hdr
            cell.border = borda_preta
            cell.alignment = alinhamento_centro
            if col in cols_estaticas:
                cell.fill = fill_preto
            elif col == 'Attribute A':
                cell.fill = fill_azul_claro
            elif col == 'Conclusion':
                cell.fill = fill_verde_claro

        # Dados
        for r, (_, row) in enumerate(df.iterrows(), start=1):
            for c, value in enumerate(row, start=0):
                cell = wss.cell(row=start_row + r, column=start_col + c, value=value)
                cell.font = font_std
                cell.border = borda_preta
                cell.alignment = alinhamento_centro
                if df.columns[c] == 'Conclusion':
                    if value == 'Effective':
                        cell.font = font_eff
                    elif value == 'Ineffective':
                        cell.font = font_ineff

    # Função para escrever dfs sem cabeçalho, com formatação específica
    def escreve_df_nohdr(wss, df, start_row, start_col):
        borda_branca = Border(left=Side(style="thin", color="FFFFFF"),
                             right=Side(style="thin", color="FFFFFF"),
                             top=Side(style="thin", color="FFFFFF"),
                             bottom=Side(style="thin", color="FFFFFF"))

        for r, (_, row) in enumerate(df.iterrows(), start=0):
            for c, value in enumerate(row, start=0):
                cell = wss.cell(row=start_row + r, column=start_col + c, value=value)
                if c == 0:  # Coluna B
                    cell.font = font_hdr
                    cell.fill = fill_preto
                    cell.border = borda_branca
                elif c == 1:  # Coluna C
                    cell.font = font_std
                    cell.border = borda_preta
                else:
                    cell.font = font_std
                    cell.border = borda_preta
                cell.alignment = alinhamento_centro

    # Abrir aba summary e limpar conteúdo a partir da linha definida
    wss = wb[aba_destino_summary]
    max_col_letter = get_column_letter(wss.max_column or 1)
    max_row = wss.max_row or linha_clear_summary
    for row in wss[f"A{linha_clear_summary}:{max_col_letter}{max_row}"]:
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()

    # Preencher infos fixas
    wss["C4"] = nome_cliente
    wss["C5"] = fy_start
    wss["C6"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    wss["C7"] = __DASversion__
    wss["C8"] = __TARAversion__

    # Escrever df_summary_attribute_a
    escreve_df_summary(wss, df_summary_attribute_a, start_row=linha_inicio_summary, start_col=coluna_inicio_summary)

    # Escrever df_extraction_date abaixo do summary_attribute_a
    linha_inicio_extraction = linha_inicio_summary + len(df_summary_attribute_a) + 4
    cell = wss.cell(row=linha_inicio_extraction - 2, column=coluna_inicio_summary,
                    value="Table 3 - Input Information: Considered Extraction Date")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FF009A44")
    cell.alignment = alinhamento_esquerda
    escreve_df_nohdr(wss, df_extraction_date, start_row=linha_inicio_extraction, start_col=coluna_inicio_summary)

    # Escrever df_quantidade abaixo
    linha_inicio_qtd = linha_inicio_extraction + len(df_extraction_date) + 3
    cell = wss.cell(row=linha_inicio_qtd - 2, column=coluna_inicio_summary,
                    value="Table 4 - Input Information: Number of Records")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FF009A44")
    cell.alignment = alinhamento_esquerda
    escreve_df_nohdr(wss, df_quantidade, start_row=linha_inicio_qtd, start_col=coluna_inicio_summary)

    # Escrever df_tipo_teste logo abaixo
    linha_inicio_teste = linha_inicio_qtd + len(df_quantidade) + 3
    cell = wss.cell(row=linha_inicio_teste - 2, column=coluna_inicio_summary,
                    value="Table 5 - Input Information: Test Type")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FF009A44")
    cell.alignment = alinhamento_esquerda
    escreve_df_nohdr(wss, df_tipo_teste, start_row=linha_inicio_teste, start_col=coluna_inicio_summary)

    # Escrever df_temp_value logo abaixo
    linha_inicio_temp = linha_inicio_teste + len(df_tipo_teste) + 3
    cell = wss.cell(row=linha_inicio_temp - 2, column=coluna_inicio_summary,
                    value="Table 6 - Input Information: Maximum Timeliness Considered")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FF009A44")
    cell.alignment = alinhamento_esquerda
    escreve_df_nohdr(wss, df_temp_value, start_row=linha_inicio_temp, start_col=coluna_inicio_summary)

    # Ajustar larguras fixas colunas
    larguras_fixas = {
        "A": 3,
        "B": 35,
        "C": 30,
        "D": 45,
        "E": 40,
        "F": 50,
        "G": 30,
        "H": 15,
    }
    for col_letra, largura in larguras_fixas.items():
        wss.column_dimensions[col_letra].width = largura


def caminho_explorer(caminho_output):
    return str(Path(caminho_output).parent)


## Funções que se diferenciam para teste apenas de A

# Def para criar tabela de teste atributo B apenas com cabeçalhos

def cabeçalho_b():
    """
    Retorna um DataFrame vazio com as colunas padrão do teste de Atributo B.
    """
    colunas = [
        "ID",
        "Nome",
        "Cargo",
        "Centro de Custo",
        "Data de Desligamento",
        "Sistema",
        "Status",
        "Data de Bloqueio",
        "Tempestividade",
        "Atributo B",
        "Conclusão"
    ]

    df = pd.DataFrame(columns=colunas)

    df.loc[0] = [None]*len(colunas)

    return df


# Escrever a aba de Summary para opção de testar só A:

def escrever_summary_a(
    wb,
    nome_cliente,
    fy_start,
    df_summary_attribute_a,
    df_extraction_date,
    df_quantidade,
    df_tipo_teste
):
    # Configurações básicas
    aba_destino_summary = "APP.02 Summary"
    linha_clear_summary = 45          # limpar a partir daqui
    linha_inicio_summary = 44         # onde começa o resumo
    coluna_inicio_summary = 2         # coluna B

    # Estilos - reutilizando os mesmos do contexto anterior
    font_hdr   = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_std   = Font(name="Calibri", size=11)
    font_eff   = Font(name="Calibri", size=11, bold=True, color="00B050")  # verde
    font_ineff = Font(name="Calibri", size=11, bold=True, color="FF0000")  # vermelho
    fill_preto       = PatternFill("solid", fgColor="000000")
    fill_azul_claro  = PatternFill("solid", fgColor="00B0F0")
    fill_verde_claro = PatternFill("solid", fgColor="92D050")
    borda_preta = Border(left=Side(style="thin", color="000000"),
                         right=Side(style="thin", color="000000"),
                         top=Side(style="thin", color="000000"),
                         bottom=Side(style="thin", color="000000"))

    alinhamento_centro = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alinhamento_esquerda = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # Função para escrever o df_summary_attribute_a (com cabeçalho e formatação)
    def escreve_df_summary(wss, df, start_row, start_col):
        cols_estaticas = ['Information', 'Value']
        
        # Cabeçalhos
        for idx, col in enumerate(df.columns):
            cell = wss.cell(row=start_row, column=start_col + idx, value=col)
            cell.font = font_hdr
            cell.border = borda_preta
            cell.alignment = alinhamento_centro
            if col in cols_estaticas:
                cell.fill = fill_preto
            elif col == 'Attribute A':
                cell.fill = fill_azul_claro
            elif col == 'Conclusion':
                cell.fill = fill_verde_claro

        # Dados
        for r, (_, row) in enumerate(df.iterrows(), start=1):
            for c, value in enumerate(row, start=0):
                cell = wss.cell(row=start_row + r, column=start_col + c, value=value)
                cell.font = font_std
                cell.border = borda_preta
                cell.alignment = alinhamento_centro
                if df.columns[c] == 'Conclusion':
                    if value == 'Effective':
                        cell.font = font_eff
                    elif value == 'Ineffective':
                        cell.font = font_ineff

    # Função para escrever dfs sem cabeçalho, com formatação específica
    def escreve_df_nohdr(wss, df, start_row, start_col):
        borda_branca = Border(left=Side(style="thin", color="FFFFFF"),
                             right=Side(style="thin", color="FFFFFF"),
                             top=Side(style="thin", color="FFFFFF"),
                             bottom=Side(style="thin", color="FFFFFF"))

        for r, (_, row) in enumerate(df.iterrows(), start=0):
            for c, value in enumerate(row, start=0):
                cell = wss.cell(row=start_row + r, column=start_col + c, value=value)
                if c == 0:  # Coluna B
                    cell.font = font_hdr
                    cell.fill = fill_preto
                    cell.border = borda_branca
                elif c == 1:  # Coluna C
                    cell.font = font_std
                    cell.border = borda_preta
                else:
                    cell.font = font_std
                    cell.border = borda_preta
                cell.alignment = alinhamento_centro

    # Abrir aba summary e limpar conteúdo a partir da linha definida
    wss = wb[aba_destino_summary]
    max_col_letter = get_column_letter(wss.max_column or 1)
    max_row = wss.max_row or linha_clear_summary
    for row in wss[f"A{linha_clear_summary}:{max_col_letter}{max_row}"]:
        for cell in row:
            cell.value = None
            cell.fill = PatternFill()
            cell.font = Font()
            cell.border = Border()

    # Preencher infos fixas
    wss["C4"] = nome_cliente
    wss["C5"] = fy_start
    wss["C6"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    wss["C7"] = __DASversion__
    wss["C8"] = __TARAversion__

    # Escrever df_summary_attribute_a
    escreve_df_summary(wss, df_summary_attribute_a, start_row=linha_inicio_summary, start_col=coluna_inicio_summary)

    # Escrever df_extraction_date abaixo do summary_attribute_a
    linha_inicio_extraction = linha_inicio_summary + len(df_summary_attribute_a) + 4
    cell = wss.cell(row=linha_inicio_extraction - 2, column=coluna_inicio_summary,
                    value="Table 3 - Input Information: Considered Extraction Date")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FF009A44")
    cell.alignment = alinhamento_esquerda
    escreve_df_nohdr(wss, df_extraction_date, start_row=linha_inicio_extraction, start_col=coluna_inicio_summary)

    # Escrever df_quantidade abaixo
    linha_inicio_qtd = linha_inicio_extraction + len(df_extraction_date) + 3
    cell = wss.cell(row=linha_inicio_qtd - 2, column=coluna_inicio_summary,
                    value="Table 4 - Input Information: Number of Records")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FF009A44")
    cell.alignment = alinhamento_esquerda
    escreve_df_nohdr(wss, df_quantidade, start_row=linha_inicio_qtd, start_col=coluna_inicio_summary)

    # Escrever df_tipo_teste logo abaixo
    linha_inicio_teste = linha_inicio_qtd + len(df_quantidade) + 3
    cell = wss.cell(row=linha_inicio_teste - 2, column=coluna_inicio_summary,
                    value="Table 5 - Input Information: Test Type")
    cell.font = Font(name="Calibri", size=11, bold=True, color="FF009A44")
    cell.alignment = alinhamento_esquerda
    escreve_df_nohdr(wss, df_tipo_teste, start_row=linha_inicio_teste, start_col=coluna_inicio_summary)


    # Ajustar larguras fixas colunas
    larguras_fixas = {
        "A": 3,
        "B": 35,
        "C": 30,
        "D": 45,
        "E": 40,
        "F": 50,
        "G": 30,
        "H": 15,
    }
    for col_letra, largura in larguras_fixas.items():
        wss.column_dimensions[col_letra].width = largura


def qtd_register_a(df1):
    # Contagem por sistema em cada df
    contagem_1 = df1['Sistema'].value_counts()
    
    # Transforma em DataFrame
    df_resultado = contagem_1.reset_index()
    df_resultado.columns = ['Sistema', 'Qtd']

    return df_resultado